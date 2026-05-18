import datetime
import openpyxl

wb = openpyxl.load_workbook('阿里巴巴2020年股票数据.xlsx')

print(wb.sheetnames)

# 获取第一个sheet
sheet = wb.worksheets[0]

# 获取单元格范围
print(sheet.dimensions)

# 获取最大行和最大列
print(sheet.max_row, sheet.max_column)

# 获取指定单元格内容
print(sheet.cell(3, 3).value)  # 0 开始

print(sheet['C3'].value)
print(sheet['G255'].value)

# 获取多个单元格（嵌套元组）
print(sheet['A2:C5'])

for row_ch in range(2, sheet.max_row + 1):
    for col_ch in 'ABCDEFG':
        value = sheet[f'{col_ch}{row_ch}'].value
        if type(value) == datetime.datetime:
            print(value.strftime('%Y年%m月%d日'), end='\t')
        elif type(value) == int:
            print(f'{value:<10d}', end='\t')
        elif type(value) == float:
            print(f'{value:.4f}',end='\t')
        else:
            print(value,end='\t')
    print()