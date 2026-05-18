"""
openpyxl的优点在于，当我们打开一个 Excel 文件后，
既可以对它进行读操作，又可以对它进行写操作，而且在操作的便捷性上是优于xlwt和xlrd的。
此外，如果要进行样式编辑和公式计算，使用openpyxl也远比上一个章节我们讲解的方式更为简单，而且openpyxl还支持数据透视和插入图表等操作，功能非常强大。
有一点需要再次强调，openpyxl并不支持操作 Office 2007 以前版本的 Excel 文件。
"""

import datetime
import openpyxl

# 加载一个工作簿
wb = openpyxl.load_workbook("阿里巴巴2020年股票数据.xlsx")
# 获取工作簿的名字
print(wb.sheetnames)
# 获取第一个工作表  --> workSheet
sheet = wb.worksheets[0]
# 获取行数和列数
print(sheet.max_row,sheet.max_column)


# 获取指定单元格数据
print(sheet.cell(3,3).value)
print(sheet['C3'].value)
print(sheet['G255'].value)

# 获取多个单元格数据
print(sheet['A2:C5'])

# 读取单元格数据
for row_ch in range(2,sheet.max_row + 1):
    for col_ch in 'ABCDEFG':
        value = sheet[f'{col_ch}{row_ch}'].value
        if type(value) == datetime.datetime:
            print(value.strftime('%Y年%m月%d日'),end="\t")
        elif type(value) == int:
            print(f'{value:<10d}',end="\t")
        elif type(value) == float:
            print(f'{value:.4f}',end="\t")
        else:
            print(value,end="\t")
    print()












